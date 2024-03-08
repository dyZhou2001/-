import openpyxl
import os
desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
def setsheet(sheetname):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.merge_cells('A1:C1')
    ws1.merge_cells('D1:F1')
    ws1.merge_cells('A2:F3')

    ws1['A4']="序号"
    ws1['B4'] = "品名"
    ws1['C4'] = "单价"
    ws1['D4'] = "数量"
    ws1['E4'] = "规格"
    ws1['F4'] = "总价"
    wb.save(desktop_path+"\\"+sheetname+".xlsx")

def completeSheet(customer):
    wb = openpyxl.load_workbook(desktop_path+"\\"+str(customer.name+customer.address)+".xlsx")
    ws=wb.active
    goodline=1
    rownum=goodline+4
    ws['A1']="客户： "+customer.name
    ws['D1'] = "电话： " + customer.phonenum
    ws['A2']=customer.address+"          \n"+customer.date
    for good in customer.goodlist:
        ws['A'+str(rownum)]=goodline
        ws['B' + str(rownum)] = good.goodnm
        ws['C' + str(rownum)] = good.goodpr
        ws['D' + str(rownum)] = good.number
        if good.danwei=="平米":
            good.danwei="m²"
        ws['E' + str(rownum)] = good.danwei
        ws['F' + str(rownum)] = good.allmoney
        goodline+=1
        rownum+=1
    ws.merge_cells('A'+str(rownum)+":"+'C'+str(rownum))
    ws.merge_cells('D'+str(rownum)+":"+'F'+str(rownum))
    ws['A' + str(rownum)] = "合计： "
    ws['D'+str(rownum)]='=SUM(F5:F'+str(rownum)+')'
    rownum+=1
    ws.merge_cells('A'+str(rownum)+":"+'F'+str(rownum))
    ws.row_dimensions[rownum].height=48
    ws['A'+str(rownum)]="备注： "
    wb.save(desktop_path + "\\" + str(customer.name+customer.address) + ".xlsx")










# sheetname="test"
# desktop_path = os.path.join(os.environ['USERPROFILE'], 'Desktop')
# wb = openpyxl.Workbook()
# ws1 = wb.active
# ws1.merge_cells('A1:C1')
# ws1.merge_cells('D1:F1')
# ws1.merge_cells('A2:F3')
# ws1.merge_cells('A11:F13')
# wb.save(desktop_path+"\\"+sheetname+".xlsx")


